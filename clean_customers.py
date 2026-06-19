#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ทำความสะอาดฐานลูกค้าจาก sheet001.html -> Excel 2 ไฟล์
  1) Customers_clean_v2_20260617.xlsx  (ฐานสะอาดพร้อมใช้)
  2) Issues_to_fix_20260617.xlsx       (รายการที่ต้องแก้)
ลบ "รหัสนำหน้าชื่อ" ทุกรูปแบบ โดยพยายามเก็บชื่อแบรนด์จริง (111, 369, G33, C-CON) ไว้ + ติดธงรายที่กำกวม
"""
from html.parser import HTMLParser
import re
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = 'sheet001.html'
OUT_CLEAN = 'Customers_clean_v2_20260617.xlsx'
OUT_ISSUES = 'Issues_to_fix_20260617.xlsx'
MY_SALE = 'Mr Teerawat Chatpomarat'

class TableParser(HTMLParser):
    def __init__(self):
        super().__init__(); self.rows=[]; self.cur=None; self.cell=None
    def handle_starttag(self,t,a):
        if t=='tr': self.cur=[]
        elif t in('td','th'): self.cell=[]
    def handle_endtag(self,t):
        if t in('td','th'):
            if self.cur is not None and self.cell is not None: self.cur.append(''.join(self.cell).strip())
            self.cell=None
        elif t=='tr':
            if self.cur is not None: self.rows.append(self.cur); self.cur=None
    def handle_data(self,d):
        if self.cell is not None: self.cell.append(d)

def load_records():
    data=open(SRC,encoding='utf-8',errors='ignore').read()
    p=TableParser(); p.feed(data)
    rows=[r for r in p.rows if any(c.strip() for c in r)]
    return [[re.sub(r'\s+',' ',c).strip() for c in (r+['']*7)[:7]] for r in rows[1:]]

# ---------- leading-code rules (run repeatedly until stable) ----------
THAI_ORG = r'บริษัท|ห้างหุ้นส่วน|หจก|บจก|นาย|นาง|นางสาว|คุณ|ร้าน|องค์การ|เทศบาล|การไฟฟ้า|อบต|อบจ|ที่ว่าการ|โรงเรียน|โรงพยาบาล|มหาวิทยาลัย|วัด|สำนักงาน'
RULES = [
    re.compile(r'^\(\s*[A-Za-z]{0,3}-?\d{2,8}\s*\)\s*'),          # (C-00001) (N6305003) (641201)
    re.compile(r'^[A-Za-z]{1,2}-\d{2,8}\s*[-.\s]+'),              # L-251  S-00259  TF-0225  L-00077-
    re.compile(r'^\d{1,3}-[A-Za-z]{1,2}-?\d+\s*[-\s]+'),          # 15-C-00498  72-L-768
    re.compile(r'^\d{1,3}-\d{2,8}\s*[-\s]+'),                     # 10-60706-  60-60701
    re.compile(r'^[A-Za-z]{1,3}\d{3,8}\s*[-.\s]+'),               # B6701 EN6808 K001- N6305003
    re.compile(r'^[A-Za-z]{1,3}\d{3,8}(?=[ก-๙])'),                # glued -> Thai
    re.compile(r'^[A-Za-z]{1,3}\d{4,8}(?=[A-Z][a-z])'),           # Tx71223Chin (glued -> Capitalized)
    re.compile(r'^\d{4,8}\s*-\s*'),                               # 71009-  6605016-
    re.compile(r'^6\d{3,7}\s*[-.\s]+'),                           # 6708  670720  6612.  650609-
    re.compile(r'^6\d{3,7}(?=[ก-๙A-Za-z])'),                      # 660824คุณ (glued)
    re.compile(r'^\d{6,7}\s+'),                                   # 170767  230167  402707 (date codes)
    re.compile(r'^\d{3,6}\s+(?=(?:%s))' % THAI_ORG),              # 0801 บริษัท  (digit code + Thai org word)
    re.compile(r'^\d{2,4}\s*-\s*(?=[^\d\s-])'),                   # 96-  6808 -
]
def strip_codes(n):
    prev=None
    while prev!=n:
        prev=n
        for rx in RULES: n=rx.sub('',n).strip()
    return n
def clean_name(n):
    return re.sub(r'\s+',' ', strip_codes(n).replace('?','').strip())
def removed_prefix(orig, nocode):
    return orig[:orig.index(nocode)].strip() if nocode and nocode in orig else ''

def taxid_status(tax):
    d=re.sub(r'\D','',tax)
    if len(d)!=13: return 'ไม่ครบ/ว่าง'
    if len(set(d))==1: return 'ผิดหลัก'
    s=sum(int(d[i])*(13-i) for i in range(12))
    return 'ถูกหลัก' if (11-(s%11))%10==int(d[12]) else 'ผิดหลัก'
def is_junk(nm):
    if not nm: return True
    if re.fullmatch(r'[\W\d_]+',nm): return True
    return nm.strip().lower() in ('test','tes','aaa','abc','na','n/a')
def ctype(tax):
    d=re.sub(r'\D','',tax)
    if len(d)<13: return 'ไม่ครบ'
    if d.startswith('0994'): return 'ราชการ'
    if d[0]=='0': return 'นิติบุคคล'
    if d[0] in '12345678': return 'บุคคล'
    return 'อื่นๆ'
def ambiguous(nm):  # ชื่อยังขึ้นต้นคล้ายรหัส -> ให้คนตรวจ
    return bool(re.match(r'^[A-Za-z]{1,3}\d{2}\b',nm) or re.match(r'^\d',nm))

def style_header(ws,head,color):
    f=PatternFill('solid',fgColor=color); ft=Font(bold=True,color='FFFFFF',size=11)
    for c in range(1,len(head)+1):
        cl=ws.cell(1,c); cl.fill=f; cl.font=ft
        cl.alignment=Alignment(vertical='center',horizontal='center',wrap_text=True)
    ws.row_dimensions[1].height=30; ws.freeze_panes='A2'

def main():
    recs=load_records()
    allnames=[clean_name(r[1]) for r in recs]
    dupset={k for k,v in Counter(n for n in allnames if n and not is_junk(n)).items() if v>1}

    # ----- clean file -----
    wb=openpyxl.Workbook(); ws=wb.active; ws.title='Customers'
    H=['ID','ชื่อลูกค้า','รหัสเดิม','Tel','Email','TaxID','TaxID ถูกหลัก','ประเภท','Sale','ที่อยู่','หมายเหตุ']
    ws.append(H); kept=removed=ncode=namb=0
    for r in recs:
        o=r[1]; nc=strip_codes(o); nm=clean_name(o)
        if is_junk(nm): removed+=1; continue
        if nm!=re.sub(r'\s+',' ',o.replace('?','').strip()): ncode+=1
        ts=taxid_status(r[4]); note=[]
        if ts=='ผิดหลัก': note.append('TaxID ผิดหลัก')
        elif ts=='ไม่ครบ/ว่าง': note.append('TaxID ไม่ครบ')
        if '?' in o: note.append('เคยมีอักขระเพี้ยน-ตรวจชื่อ')
        if ambiguous(nm): note.append('ตรวจชื่อ-อาจมีรหัส/แบรนด์'); namb+=1
        ws.append([r[0],nm,removed_prefix(o,nc),r[2],r[3],r[4],
                   '✓' if ts=='ถูกหลัก' else '',ctype(r[4]),r[5],r[6],' · '.join(note)]); kept+=1
    style_header(ws,H,'0F766E')
    for i,w in enumerate([7,40,12,15,26,16,12,11,22,46,28],1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.auto_filter.ref=f"A1:{get_column_letter(len(H))}{ws.max_row}"
    mine=PatternFill('solid',fgColor='E6F5F3'); amb=PatternFill('solid',fgColor='FEF3E2')
    for row in range(2,ws.max_row+1):
        ws.cell(row,2).font=Font(bold=True)
        if ws.cell(row,9).value==MY_SALE:
            for c in range(1,len(H)+1): ws.cell(row,c).fill=mine
        if 'ตรวจชื่อ-อาจ' in (ws.cell(row,11).value or ''): ws.cell(row,11).fill=amb
    wb.save(OUT_CLEAN)

    # ----- issues file -----
    wb2=openpyxl.Workbook(); w2=wb2.active; w2.title='ต้องแก้'
    H2=['ID','ชื่อ (ล้างแล้ว)','ชื่อเดิม','TaxID','สถานะ TaxID','Tel','Email','Sale','ปัญหาที่พบ','ความสำคัญ']
    w2.append(H2); nhi=nlo=0
    for r in recs:
        o=r[1]; nm=clean_name(o); ts=taxid_status(r[4]); pr=[]; hi=False
        if ts=='ผิดหลัก': pr.append('TaxID ผิดหลัก'); hi=True
        if '?' in o: pr.append('ชื่อมีอักขระเพี้ยน(?)'); hi=True
        if is_junk(nm): pr.append('ชื่อขยะ/ไม่ใช่ชื่อจริง'); hi=True
        if ambiguous(nm): pr.append('อาจมีรหัส/แบรนด์ตกค้าง'); hi=True
        em=r[3].lower()
        if em in('test@test.com','a@a.com') or em.startswith('test@'): pr.append('อีเมลปลอม')
        if ts=='ไม่ครบ/ว่าง': pr.append('TaxID ไม่ครบ/ว่าง')
        if nm and nm in dupset: pr.append('ชื่อซ้ำ')
        if not pr: continue
        sev='สูง' if hi else 'ต่ำ'; nhi+=hi; nlo+=(not hi)
        w2.append([r[0],nm,o,r[4],ts,r[2],r[3],r[5],' · '.join(pr),sev])
    style_header(w2,H2,'B45309')
    for i,w in enumerate([7,34,34,16,13,14,24,22,42,11],1): w2.column_dimensions[get_column_letter(i)].width=w
    w2.auto_filter.ref=f"A1:{get_column_letter(len(H2))}{w2.max_row}"
    hf=PatternFill('solid',fgColor='FDEAEA')
    for row in range(2,w2.max_row+1):
        if w2.cell(row,10).value=='สูง':
            w2.cell(row,10).font=Font(color='DC2626',bold=True); w2.cell(row,9).fill=hf
    wb2.save(OUT_ISSUES)

    # ----- verify -----
    names=[clean_name(r[1]) for r in recs if not is_junk(clean_name(r[1]))]
    detect=re.compile(r'^(\(?\s*[A-Za-z]{0,3}-?\d{3,}|\d{1,3}-\d|\d{4,8}[\s-]|6\d{3,7}|[A-Za-z]{1,3}\d{3,})')
    left=[n for n in names if detect.match(n)]
    print(f"ลบรหัส: {ncode} · เก็บ {kept} · ตัดขยะ {removed} · ติดธงตรวจชื่อ {namb}")
    print(f"รายงานต้องแก้: {nhi+nlo} (สูง {nhi} · ต่ำ {nlo})")
    print(f"รหัสตกค้าง (ตรวจกว้าง): {len(left)}")
    for n in left: print("   »",repr(n))

if __name__=='__main__':
    main()
